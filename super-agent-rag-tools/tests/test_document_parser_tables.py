import base64
import json
import unittest
from io import BytesIO

from rag_tools.document_parser import parse_document
from rag_tools.schemas.document_parse import DocumentParseRequest


class DocumentParserTableSampleTest(unittest.TestCase):
    def test_parse_xlsx_table_rows_and_cell_metadata(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metrics"
        sheet["B2"] = "Product"
        sheet["C2"] = "Sales"
        sheet["D2"] = "Amount"
        sheet["B3"] = "A"
        sheet["C3"] = 10
        sheet["D3"] = 100
        sheet["B4"] = "B"
        sheet["C4"] = 20
        sheet["D4"] = 300
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        response = self._parse_bytes("sample.xlsx", "XLSX", buffer.getvalue())

        table = self._single_table(response.blocks)
        self.assertEqual(
            [
                ["Product", "Sales", "Amount"],
                ["A", "10", "100"],
                ["B", "20", "300"],
            ],
            table.table_rows,
        )

        metadata = json.loads(table.metadata_json)
        self.assertIn("工作表：Metrics", metadata.get("tableContextText", ""))
        self.assertIn("tableContext", table.content_with_weight)
        cells = metadata.get("tableCellMetadata") or []
        self.assertEqual(9, len(cells))
        first_cell = self._find_cell(cells, 1, 1)
        self.assertEqual("Product", first_cell.get("value"))
        self.assertEqual(2, first_cell.get("sourceRowNo"))
        self.assertEqual(2, first_cell.get("sourceColumnNo"))
        self.assertEqual("B2", first_cell.get("excelAddress"))
        self.assertEqual("B2", first_cell.get("cellCoordinate"))
        self.assertEqual("Metrics", first_cell.get("sheetName"))

    def test_parse_xlsx_merged_title_row_keeps_metadata_and_uses_real_header(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Merged"
        sheet.merge_cells("B2:D2")
        sheet["B2"] = "Quarter Metrics"
        sheet["B3"] = "Product"
        sheet["C3"] = "Sales"
        sheet["D3"] = "Amount"
        sheet["B4"] = "A"
        sheet["C4"] = 10
        sheet["D4"] = 100
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        response = self._parse_bytes("merged.xlsx", "XLSX", buffer.getvalue())

        table = self._single_table(response.blocks)
        self.assertEqual(
            [
                ["Product", "Sales", "Amount"],
                ["A", "10", "100"],
            ],
            table.table_rows,
        )
        self.assertIn("Quarter Metrics", table.text)

        metadata = json.loads(table.metadata_json)
        self.assertEqual(["Quarter Metrics"], metadata.get("tableTitleRows"))
        title_cells = metadata.get("tableTitleCellMetadata") or []
        title_anchor = self._find_cell(title_cells, 1, 1)
        self.assertEqual("B2:D2", title_anchor.get("mergedCellRange"))
        self.assertEqual("B2", title_anchor.get("mergedCellAnchorAddress"))
        self.assertEqual(3, title_anchor.get("columnSpan"))
        self.assertTrue(title_anchor.get("mergedCellAnchor"))

        cells = metadata.get("tableCellMetadata") or []
        header_cell = self._find_cell(cells, 1, 1)
        self.assertEqual("Product", header_cell.get("value"))
        self.assertEqual(3, header_cell.get("sourceRowNo"))
        self.assertEqual("B3", header_cell.get("excelAddress"))

    def test_parse_xlsx_multi_level_header_is_flattened(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Multi"
        sheet["B2"] = "Product"
        sheet.merge_cells("C2:D2")
        sheet["C2"] = "Q1"
        sheet.merge_cells("E2:F2")
        sheet["E2"] = "Q2"
        sheet["C3"] = "Sales"
        sheet["D3"] = "Amount"
        sheet["E3"] = "Sales"
        sheet["F3"] = "Amount"
        sheet["B4"] = "A"
        sheet["C4"] = 10
        sheet["D4"] = 100
        sheet["E4"] = 20
        sheet["F4"] = 300
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        response = self._parse_bytes("multi.xlsx", "XLSX", buffer.getvalue())

        table = self._single_table(response.blocks)
        self.assertEqual(
            [
                ["Product", "Q1 Sales", "Q1 Amount", "Q2 Sales", "Q2 Amount"],
                ["A", "10", "100", "20", "300"],
            ],
            table.table_rows,
        )

        metadata = json.loads(table.metadata_json)
        self.assertTrue(metadata.get("tableHeaderFlattened"))
        self.assertEqual(
            [
                ["Product", "Q1", "", "Q2", ""],
                ["", "Sales", "Amount", "Sales", "Amount"],
            ],
            metadata.get("tableHeaderRows"),
        )
        cells = metadata.get("tableCellMetadata") or []
        q1_amount_header = self._find_cell(cells, 1, 3)
        self.assertEqual("Q1 Amount", q1_amount_header.get("value"))
        self.assertEqual("D3", q1_amount_header.get("excelAddress"))
        data_cell = self._find_cell(cells, 2, 3)
        self.assertEqual("100", data_cell.get("value"))
        self.assertEqual(4, data_cell.get("sourceRowNo"))
        self.assertEqual("D4", data_cell.get("excelAddress"))

    def test_parse_xlsx_vertical_merged_data_cells_are_filled(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "VerticalMerge"
        sheet["B2"] = "Product"
        sheet["C2"] = "Region"
        sheet["D2"] = "Amount"
        sheet["B3"] = "A"
        sheet.merge_cells("C3:C4")
        sheet["C3"] = "North"
        sheet["D3"] = 100
        sheet["B4"] = "B"
        sheet["D4"] = 120
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()

        response = self._parse_bytes("vertical-merged.xlsx", "XLSX", buffer.getvalue())

        table = self._single_table(response.blocks)
        self.assertEqual(
            [
                ["Product", "Region", "Amount"],
                ["A", "North", "100"],
                ["B", "North", "120"],
            ],
            table.table_rows,
        )

        metadata = json.loads(table.metadata_json)
        cells = metadata.get("tableCellMetadata") or []
        filled_cell = self._find_cell(cells, 3, 2)
        self.assertEqual("North", filled_cell.get("value"))
        self.assertEqual("C4", filled_cell.get("excelAddress"))
        self.assertEqual("C3:C4", filled_cell.get("mergedCellRange"))
        self.assertTrue(filled_cell.get("mergedValueFilled"))

    def test_parse_pdf_table_rows_and_cell_bboxes(self) -> None:
        try:
            import fitz
        except ImportError:
            self.skipTest("pymupdf is not installed")
        if not hasattr(fitz.Page, "find_tables"):
            self.skipTest("pymupdf table detection is not available")

        response = self._parse_bytes("sample.pdf", "PDF", self._pdf_table_bytes(fitz))

        table = self._single_table(response.blocks)
        self.assertEqual(
            [
                ["Product", "Sales", "Amount"],
                ["A", "10", "100"],
                ["B", "20", "300"],
            ],
            table.table_rows,
        )
        self.assertEqual(1, table.page_no)

        metadata = json.loads(table.metadata_json)
        cells = metadata.get("tableCellMetadata") or []
        self.assertEqual(9, len(cells))
        second_header = self._find_cell(cells, 1, 2)
        self.assertEqual("Sales", second_header.get("value"))
        self.assertEqual("R1C2", second_header.get("cellCoordinate"))
        self.assertEqual(1, second_header.get("pageNo"))
        bbox = json.loads(second_header.get("bboxJson") or "{}")
        self.assertAlmostEqual(192.0, bbox.get("x0"), places=1)
        self.assertAlmostEqual(120.0, bbox.get("y0"), places=1)
        self.assertAlmostEqual(292.0, bbox.get("x1"), places=1)
        self.assertAlmostEqual(152.0, bbox.get("y1"), places=1)

    def _parse_bytes(self, file_name: str, file_type: str, content: bytes):
        return parse_document(
            DocumentParseRequest(
                fileName=file_name,
                fileType=file_type,
                contentBase64=base64.b64encode(content).decode("ascii"),
            )
        )

    def _single_table(self, blocks):
        tables = [block for block in blocks if block.block_type == "TABLE"]
        self.assertEqual(1, len(tables))
        return tables[0]

    def _find_cell(self, cells: list[dict], row_no: int, column_no: int) -> dict:
        for cell in cells:
            if cell.get("rowNo") == row_no and cell.get("columnNo") == column_no:
                return cell
        self.fail(f"Missing cell R{row_no}C{column_no}")

    def _pdf_table_bytes(self, fitz) -> bytes:
        document = fitz.open()
        page = document.new_page(width=595, height=842)
        x0 = 72
        y0 = 120
        column_widths = [120, 100, 100]
        row_height = 32
        rows = [
            ["Product", "Sales", "Amount"],
            ["A", "10", "100"],
            ["B", "20", "300"],
        ]
        for row_index in range(len(rows) + 1):
            y = y0 + row_index * row_height
            page.draw_line((x0, y), (x0 + sum(column_widths), y), color=(0, 0, 0), width=0.8)
        current_x = x0
        for column_index in range(len(column_widths) + 1):
            page.draw_line((current_x, y0), (current_x, y0 + len(rows) * row_height), color=(0, 0, 0), width=0.8)
            if column_index < len(column_widths):
                current_x += column_widths[column_index]
        for row_index, row in enumerate(rows):
            current_x = x0
            for column_index, value in enumerate(row):
                page.insert_text((current_x + 6, y0 + row_index * row_height + 21), value, fontsize=11)
                current_x += column_widths[column_index]
        buffer = BytesIO()
        document.save(buffer)
        document.close()
        return buffer.getvalue()


if __name__ == "__main__":
    unittest.main()
